import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime
import boto3

iam = boto3.client("iam")

users = iam.list_users()

# Risk Classification Logic
def classify_risk(username, policies, mfa_enabled):
    risk_level = "LOW"
    reasons = []

    for policy in policies:
        # Check 1 - Admin access without MFA = CRITICAL
        if policy == "AdministratorAccess":
            if not mfa_enabled:
                risk_level = "CRITICAL"
                reasons.append("Admin access without MFA")
            else:
                risk_level = "HIGH"
                reasons.append("Has AdministratorAccess")

        # Check 2 - Intern with FullAccess = MEDIUM
        if policy == "AmazonS3FullAccess" and "intern" in username.lower():
            if risk_level not in ["CRITICAL", "HIGH"]:
                risk_level = "MEDIUM"
            reasons.append("Intern has S3 Full Access - violates least privilege")

    # Check 3 - Inactive user with access = HIGH
    if "inactive" in username.lower():
        if risk_level not in ["CRITICAL"]:
            risk_level = "HIGH"
        reasons.append("Inactive user still has active access")

    # Check 4 - Missing MFA on any account
    if not mfa_enabled:
        if risk_level == "LOW":
            risk_level = "MEDIUM"
        reasons.append("MFA not enabled")

    return risk_level, reasons
def generate_report(audit_results):
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IAM Audit Report"

    # Define colors for risk levels
    colors = {
        "CRITICAL" : "FF0000",  # Red
        "HIGH"     : "FF6600",  # Orange
        "MEDIUM"   : "FFFF00",  # Yellow
        "LOW"      : "00FF00",  # Green
    }

    # Create header row
    headers = ["Username", "Policies", "MFA Status", "Risk Level", "Reasons"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Fill data rows
    for row, result in enumerate(audit_results, start=2):
        ws.cell(row=row, column=1, value=result["username"])
        ws.cell(row=row, column=2, value=result["policies"])
        ws.cell(row=row, column=3, value=result["mfa"])
        ws.cell(row=row, column=4, value=result["risk"])
        ws.cell(row=row, column=5, value=result["reasons"])

        # Color the row based on risk level
        fill_color = colors.get(result["risk"], "FFFFFF")
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = fill

    # Auto adjust column widths
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = max_length + 5

    # Save report with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"IAM_Audit_Report_{timestamp}.xlsx"
    wb.save(filename)
    print(f"\n✅ Report saved as: {filename}")

print("\n===== CLOUD ACCESS PRIVILEGE AUDITOR =====\n")

audit_results = []

for user in users["Users"]:
    username = user["UserName"]

    # Get policies
    policies_response = iam.list_attached_user_policies(UserName=username)
    policy_names = [p["PolicyName"] for p in policies_response["AttachedPolicies"]]

    # Get MFA status
    mfa_response = iam.list_mfa_devices(UserName=username)
    mfa_enabled = len(mfa_response["MFADevices"]) > 0

    # Run risk classification
    risk, reasons = classify_risk(username, policy_names, mfa_enabled)

    # Print to screen
    print("=" * 55)
    print(f"  User     : {username}")
    print(f"  Policies : {', '.join(policy_names) if policy_names else 'None'}")
    print(f"  MFA      : {'Enabled' if mfa_enabled else 'Disabled'}")
    print(f"  Risk     : {risk}")
    print(f"  Reasons  : {', '.join(reasons) if reasons else 'No issues found'}")
    print("=" * 55)
    print()

    # Collect for report
    audit_results.append({
        "username" : username,
        "policies" : ", ".join(policy_names) if policy_names else "None",
        "mfa"      : "Enabled" if mfa_enabled else "Disabled",
        "risk"     : risk,
        "reasons"  : ", ".join(reasons) if reasons else "No issues found"
    })

# Generate Excel report
generate_report(audit_results)